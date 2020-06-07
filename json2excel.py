# -*- coding: utf-8 -*-
import openpyxl
import json
import sys

def main():
    """
    メイン処理
    """
    # 設定ファイル読み込み
    settings = load_settings()

    json_data = load_data(settings)
    book = openpyxl.load_workbook(settings['outputFile'])
    sheet = book[settings['outputSheet']]

    # ヘッダー設定
    headers = [key for key in json_data[0].keys()]
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=i+1).value = header

    # データ書き込み
    for i, data in enumerate(json_data):
        target_row = i+2
        for j, header in enumerate(headers):
            sheet.cell(row=target_row, column=j+1).value = data[header]

    book.save(settings['outputFile'])
    sys.exit()

def load_settings():
    with open('settings.json', 'r') as f:
        try:
            settings = json.load(f)

            if 'dataFile' not in settings:
                raise Exception('dataFileが設定されていません。')
            if 'outputFile' not in settings:
                raise Exception('outputFileが設定されていません。')
            if 'outputSheet' not in settings:
                raise Exception('outputSheetが設定されていません。')

        except Exception as e:
            print('設定ファイルの読み込みに失敗しました。')
            print(e)
            sys.exit(1)
    return settings

def load_data(settings):
    with open(settings['dataFile'], 'r') as f:
        try:
            json_data = json.load(f)
        except Exception as e:
            print('dataFileの読み込みに失敗しました。')
            print(e)
            sys.exit(1)
    return json_data

main()
